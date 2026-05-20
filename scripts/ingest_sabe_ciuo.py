"""
Ingestión del clasificador CIUO-08 SABE (SENCE).
Fuente: https://sence.gob.cl/sites/default/files/clasificador_ciuo08cl_sabe.xlsx
Descarga el archivo si no existe localmente y hace upsert en public.sabe_ciuo_mapping.

Uso:
  python scripts/ingest_sabe_ciuo.py              # descarga + inserta
  python scripts/ingest_sabe_ciuo.py --dry-run    # solo muestra, no inserta
"""

import sys, os, re, unicodedata
sys.stdout.reconfigure(encoding="utf-8")

import requests
import openpyxl
from supabase import create_client

# ── Configuración ──────────────────────────────────────────────────────────────
SUPABASE_URL = "https://fdogahabyuauglkfhzkj.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZkb2dhaGFieXVhdWdsa2ZoemtqIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3ODE1MTYzNSwiZXhwIjoyMDkzNzI3NjM1fQ.qewn_gGCdjBAaJ-PEy0FQKZkdgQ28VyUp3SxcQfl1sk"
SABE_URL     = "https://sence.gob.cl/sites/default/files/clasificador_ciuo08cl_sabe.xlsx"
LOCAL_FILE   = os.path.join(os.path.dirname(__file__), "data", "sabe_ciuo08.xlsx")
DRY_RUN      = "--dry-run" in sys.argv
BATCH_SIZE   = 500

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


def normalizar(texto: str) -> str:
    """Minúsculas sin tildes ni caracteres especiales."""
    texto = texto.lower().strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto


def descargar_si_necesario():
    if os.path.exists(LOCAL_FILE):
        print(f"  Usando archivo local: {LOCAL_FILE}")
        return
    print(f"  Descargando clasificador SABE desde SENCE…")
    r = requests.get(SABE_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    os.makedirs(os.path.dirname(LOCAL_FILE), exist_ok=True)
    with open(LOCAL_FILE, "wb") as f:
        f.write(r.content)
    print(f"  Guardado: {len(r.content):,} bytes")


def leer_clasificador() -> list[dict]:
    wb = openpyxl.load_workbook(LOCAL_FILE, read_only=True, data_only=True)

    # Hoja 'cargos incluidos': (ciuo4, cargo_titulo, nombre_ciuo4)
    ws_cargos = wb["cargos incluidos"]
    rows = list(ws_cargos.iter_rows(min_row=2, values_only=True))

    # Hoja 'definicion': (oficio1, nombre_oficio1, oficio4, nombre_oficio4, definicion)
    ws_def = wb["definicion"]
    definiciones: dict[int, str] = {}
    for row in ws_def.iter_rows(min_row=2, values_only=True):
        if row[2] is not None:
            definiciones[int(row[2])] = str(row[3] or "")

    registros = []
    vistos: set[tuple] = set()

    for row in rows:
        ciuo4_raw, cargo_raw, nombre_raw = row[0], row[1], row[2]
        if ciuo4_raw is None or cargo_raw is None:
            continue
        ciuo4        = int(ciuo4_raw)
        cargo_titulo = str(cargo_raw).strip()
        nombre_ciuo4 = str(nombre_raw or definiciones.get(ciuo4, "")).strip()
        cargo_norm   = normalizar(cargo_titulo)

        llave = (ciuo4, cargo_norm)
        if llave in vistos:
            continue
        vistos.add(llave)

        registros.append({
            "ciuo4":        ciuo4,
            "cargo_titulo": cargo_titulo,
            "cargo_norm":   cargo_norm,   # versión normalizada para búsquedas fuzzy
            "nombre_ciuo4": nombre_ciuo4,
        })

    print(f"  Filas únicas cargos incluidos: {len(registros):,}")
    return registros


def upsert_en_supabase(registros: list[dict]):
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    total = 0
    for i in range(0, len(registros), BATCH_SIZE):
        batch = registros[i:i + BATCH_SIZE]
        sb.table("sabe_ciuo_mapping").upsert(
            batch, on_conflict="ciuo4,cargo_norm"
        ).execute()
        total += len(batch)
        print(f"  Upsert {total}/{len(registros)}")
    print(f"  ✓ {total} registros en sabe_ciuo_mapping")


def buscar_ciuo(cargo: str, registros: list[dict]) -> int | None:
    """Búsqueda local: retorna ciuo4 del mejor match para un título de cargo."""
    cargo_n = normalizar(cargo)
    # Coincidencia exacta
    for r in registros:
        if r["cargo_norm"] == cargo_n:
            return r["ciuo4"]
    # Coincidencia parcial: el título SABE está dentro del cargo buscado
    for r in registros:
        if r["cargo_norm"] in cargo_n or cargo_n in r["cargo_norm"]:
            return r["ciuo4"]
    return None


if __name__ == "__main__":
    print("=== Ingestión Clasificador SABE CIUO-08 ===")
    descargar_si_necesario()
    registros = leer_clasificador()

    # Mostrar ejemplos gastronomía
    gastro_kws = ["garzon", "bartender", "cocinero", "chef", "mozo", "auxiliar de aseo"]
    print("\nEjemplos gastronomía:")
    for kw in gastro_kws:
        ciuo = buscar_ciuo(kw, registros)
        print(f"  '{kw}' → CIUO {ciuo}")

    if DRY_RUN:
        print("\n[DRY-RUN] No se insertó nada.")
    else:
        print(f"\nInsertando {len(registros):,} registros…")
        upsert_en_supabase(registros)

    print("\nDone.")
